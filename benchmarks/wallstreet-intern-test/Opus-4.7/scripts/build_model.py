"""Build the three-statement model as model/coco_model.xlsx.

Tabs (in order):
  Cover         - title, scenario selector cell `Scenario`, key live outputs
  Assumptions   - bear/base/bull driver inputs that flow into model via CHOOSE()
  History       - actual financials FY19-FY25 from extracted/annual.csv + segment_history.csv
  IS            - forecasted income statement, FY25A and FY26-FY28E
  BS            - forecasted balance sheet, ratio-driven
  CF            - forecasted cash flow with NWC change driven by Δrevenue
  Valuation     - DCF / exit-multiple / P/E triangulation, 12-month price target
"""
from __future__ import annotations

import csv
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

REPO = pathlib.Path(__file__).resolve().parent.parent
OUT = REPO / "model" / "coco_model.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
BODY_FONT = Font(name="Calibri", size=10)


def cw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def hdr(ws, row, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def write(ws, r, vals, *, bold=False, fill=None, fmt=None, label_only=False):
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=j, value=v)
        cell.font = BOLD_FONT if bold else BODY_FONT
        if fill is not None:
            cell.fill = fill
        if fmt and j > 1 and not label_only:
            cell.number_format = fmt


# Anchors hard-coded from FY2025 actuals (XBRL + press release)
FY25 = {
    "rev_total": 609780, "gp_americas": 187309, "gp_intl": 35286, "gp_total": 222595,
    "sga": 140063, "op_inc": 82532, "non_op": 10439, "pretax": 92971, "tax": 21651, "ni": 71320,
    "da": 1072, "sbc": 10843, "shares_dil": 59.967691, "eps_dil": 1.19,
    "cash": 196873, "ar": 81514, "inv": 111468, "tca": 421440, "ta": 461158,
    "ap": 25464, "tcl": 116432, "tl": 129616, "se": 331542,
    "cfo": 47174, "capex": 8149, "fcf": 47174 - 8149,
    # segment sales by brand
    "am_vc": 424319, "am_pl": 62731, "am_oth": 21723,
    "intl_vc": 71943, "intl_pl": 25951, "intl_oth": 3113,
}


def load_history():
    annual = list(csv.DictReader((REPO / "extracted" / "annual.csv").open(encoding="utf-8")))
    annual.sort(key=lambda r: r["fy"])
    seg = {}
    for r in csv.DictReader((REPO / "extracted" / "segment_history.csv").open(encoding="utf-8")):
        if r["metric"] == "net_sales_ytd" and r["period_end"].endswith("-12-31") and r["fp"].startswith("Q4"):
            seg[(r["period_end"][:4], r["segment"], r["brand"])] = int(r["value"])
    return annual, seg


# ---------------------------------------------------------------------------
# Assumptions
# ---------------------------------------------------------------------------
GROWTH_DRIVERS = [
    # Each row: label, FY26 (b/ba/bu), FY27 (b/ba/bu), FY28 (b/ba/bu)
    ("Americas - Vita Coco Coconut Water (% YoY)",     0.06, 0.12, 0.18,    0.06, 0.10, 0.15,    0.05, 0.09, 0.13),
    ("Americas - Private Label (% YoY)",                0.05, 0.22, 0.35,    0.04, 0.10, 0.18,    0.03, 0.07, 0.12),
    ("Americas - Other (% YoY)",                       -0.10, 0.10, 0.30,    0.05, 0.10, 0.20,    0.05, 0.08, 0.15),
    ("International - Vita Coco Coconut Water (% YoY)", 0.18, 0.30, 0.40,    0.15, 0.25, 0.35,    0.12, 0.22, 0.30),
    ("International - Private Label (% YoY)",           0.05, 0.15, 0.25,    0.05, 0.12, 0.20,    0.05, 0.10, 0.15),
    ("International - Other (% YoY)",                  -0.20, 0.00, 0.20,    0.00, 0.05, 0.15,    0.00, 0.05, 0.10),
]
GM_DRIVERS = [
    ("Americas gross margin (%)",        0.355, 0.385, 0.400,    0.355, 0.390, 0.410,    0.355, 0.390, 0.415),
    ("International gross margin (%)",   0.330, 0.360, 0.385,    0.335, 0.370, 0.400,    0.340, 0.375, 0.405),
]
SGA_DRIVERS = [
    ("SG&A as % of net sales",                  0.235, 0.220, 0.210,    0.230, 0.215, 0.205,    0.225, 0.210, 0.200),
    ("Effective tax rate (%)",                  0.250, 0.230, 0.215,    0.250, 0.230, 0.215,    0.250, 0.230, 0.215),
    ("D&A as % of net sales",                   0.0020, 0.0020, 0.0020, 0.0020, 0.0020, 0.0020, 0.0020, 0.0020, 0.0020),
    ("Stock-based comp as % of net sales",      0.018, 0.018, 0.018,    0.018, 0.018, 0.018,    0.018, 0.018, 0.018),
    ("Capex as % of net sales",                 0.0020, 0.0020, 0.0020, 0.0020, 0.0020, 0.0020, 0.0020, 0.0020, 0.0020),
    ("Net working capital as % of Delta rev",   0.20, 0.20, 0.20,        0.20, 0.20, 0.20,       0.20, 0.20, 0.20),
    ("Diluted shares outstanding (M)",          60.0, 60.0, 60.0,        60.5, 60.0, 59.5,       61.0, 60.0, 59.0),
    ("Non-operating income ($M)",               0.0, 5.0, 10.0,           0.0, 5.0, 10.0,         0.0, 5.0, 10.0),
]
VAL_DRIVERS = [
    ("Exit EV/EBITDA multiple (FY27)",  12.0, 17.0, 22.0,    12.0, 17.0, 22.0,    12.0, 17.0, 22.0),
    ("WACC (%)",                         0.10, 0.09, 0.08,   0.10, 0.09, 0.08,    0.10, 0.09, 0.08),
    ("Terminal growth rate (%)",         0.020, 0.025, 0.030, 0.020, 0.025, 0.030, 0.020, 0.025, 0.030),
    ("Cash at FY25 ($M)",                197.0, 197.0, 197.0, 197.0, 197.0, 197.0, 197.0, 197.0, 197.0),
    ("Debt at FY25 ($M)",                0.0, 0.0, 0.0,       0.0, 0.0, 0.0,       0.0, 0.0, 0.0),
]


def build_assumptions(wb: Workbook) -> dict:
    ws = wb.create_sheet("Assumptions")
    cw(ws, [44, 10, 10, 10, 10, 10, 10, 10, 10, 10, 12, 12, 12])
    write(ws, 1, [
        "Driver",
        "FY26 Bear", "FY26 Base", "FY26 Bull",
        "FY27 Bear", "FY27 Base", "FY27 Bull",
        "FY28 Bear", "FY28 Base", "FY28 Bull",
        "Live FY26", "Live FY27", "Live FY28",
    ], bold=True, fill=HEADER_FILL)
    hdr(ws, 1, 13)

    rowmap = {}

    def emit_section(title, drivers, start_row, num_fmt_picker):
        write(ws, start_row, [title], bold=True, fill=SUBHEADER_FILL)
        row = start_row + 1
        for d in drivers:
            label = d[0]
            ws.cell(row=row, column=1, value=label).font = BODY_FONT
            for ci, v in enumerate(d[1:], start=2):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = INPUT_FILL
                c.number_format = num_fmt_picker(label)
            for live_col, scen_cols in [(11, (2, 3, 4)), (12, (5, 6, 7)), (13, (8, 9, 10))]:
                sb, sba, sbu = scen_cols
                f = f"=CHOOSE(Scenario,{get_column_letter(sb)}{row},{get_column_letter(sba)}{row},{get_column_letter(sbu)}{row})"
                cc = ws.cell(row=row, column=live_col, value=f)
                cc.font = BOLD_FONT
                cc.number_format = num_fmt_picker(label)
            rowmap[label] = row
            row += 1
        return row + 1  # leave a blank row

    def pct_fmt(label):
        if "$M" in label or "(M)" in label or label.startswith("Exit EV"):
            return "#,##0.0\"x\"" if label.startswith("Exit") else "#,##0.0"
        return "0.0%"

    next_row = emit_section("REVENUE GROWTH", GROWTH_DRIVERS, 3, pct_fmt)
    next_row = emit_section("GROSS MARGIN", GM_DRIVERS, next_row, pct_fmt)
    next_row = emit_section("OPEX & MISC", SGA_DRIVERS, next_row, pct_fmt)
    next_row = emit_section("VALUATION", VAL_DRIVERS, next_row, pct_fmt)
    return rowmap


# ---------------------------------------------------------------------------
# Cover
# ---------------------------------------------------------------------------
def build_cover(wb: Workbook) -> None:
    ws = wb.create_sheet("Cover", 0)
    cw(ws, [40, 22, 18, 18, 18, 18])
    ws["A1"] = "The Vita Coco Company, Inc. (NASDAQ: COCO) - 3-Statement Model"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws["A2"] = "Built 2026-04-27. Source: FY2025 10-K (accn 0001482981-26-000022) and Q3-2025 10-Q + earnings press releases."
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="666666")

    ws["A4"] = "Scenario (1=Bear, 2=Base, 3=Bull):"
    ws["A4"].font = BOLD_FONT
    ws["B4"] = 2
    ws["B4"].fill = INPUT_FILL
    ws["B4"].font = Font(size=12, bold=True)
    ws["B4"].alignment = Alignment(horizontal="center")

    dn = DefinedName(name="Scenario", attr_text="Cover!$B$4")
    wb.defined_names["Scenario"] = dn

    ws["A6"] = "Key live outputs (recompute when scenario changes):"
    ws["A6"].font = BOLD_FONT
    write(ws, 7, ["", "FY2025A", "FY2026E", "FY2027E", "FY2028E"], bold=True, fill=SUBHEADER_FILL)

    def link(out_cell, label, formulas, fmt):
        ws[out_cell] = label
        ws[out_cell].font = BODY_FONT
        for i, f in enumerate(formulas):
            col = get_column_letter(2 + i)
            row = int(out_cell[1:])
            c = ws.cell(row=row, column=2 + i, value=f)
            c.number_format = fmt
            c.font = BODY_FONT

    # Net sales total is at IS!C8 (B8=FY25A, C8=FY26E, D8=FY27E, E8=FY28E)
    link("A8", "Net sales ($M)", ["=IS!B8/1000", "=IS!C8/1000", "=IS!D8/1000", "=IS!E8/1000"], "#,##0.0")
    link("A9", "YoY growth", ["", "=IS!C8/IS!B8-1", "=IS!D8/IS!C8-1", "=IS!E8/IS!D8-1"], "0.0%")
    link("A10", "Gross margin", ["=IS!B14/IS!B8", "=IS!C14/IS!C8", "=IS!D14/IS!D8", "=IS!E14/IS!E8"], "0.0%")
    link("A11", "Operating margin", ["=IS!B18/IS!B8", "=IS!C18/IS!C8", "=IS!D18/IS!D8", "=IS!E18/IS!E8"], "0.0%")
    link("A12", "Adj EBITDA ($M)", ["=IS!B26/1000", "=IS!C26/1000", "=IS!D26/1000", "=IS!E26/1000"], "#,##0.0")
    link("A13", "Adj EBITDA margin", ["=IS!B26/IS!B8", "=IS!C26/IS!C8", "=IS!D26/IS!D8", "=IS!E26/IS!E8"], "0.0%")
    link("A14", "Net income ($M)", ["=IS!B23/1000", "=IS!C23/1000", "=IS!D23/1000", "=IS!E23/1000"], "#,##0.0")
    link("A15", "Diluted EPS ($)", ["=IS!B31", "=IS!C31", "=IS!D31", "=IS!E31"], "$0.00")
    link("A16", "FCF ($M)", ["=CF!B8/1000", "=CF!C8/1000", "=CF!D8/1000", "=CF!E8/1000"], "#,##0.0")

    ws["A18"] = "Valuation (live):"
    ws["A18"].font = BOLD_FONT
    ws["A19"] = "12-month price target ($)"
    ws["B19"] = "=Valuation!B22"
    ws["B19"].number_format = "$0.00"
    ws["B19"].font = Font(size=12, bold=True, color="2F5496")
    ws["B19"].fill = TOTAL_FILL
    ws["A20"] = "Spot price 2026-04-27 ($)"
    ws["B20"] = 52.39
    ws["B20"].number_format = "$0.00"
    ws["A21"] = "Implied upside / (downside)"
    ws["B21"] = "=B19/B20-1"
    ws["B21"].number_format = "0.0%"
    ws["B21"].font = BOLD_FONT


# ---------------------------------------------------------------------------
# History
# ---------------------------------------------------------------------------
def build_history(wb: Workbook, annual_rows, fy_seg) -> None:
    ws = wb.create_sheet("History")
    cw(ws, [42, 14, 14, 14, 14, 14, 14, 14])
    headers = ["($000s unless noted)"] + [f"FY{r['fy']}A" for r in annual_rows]
    write(ws, 1, headers, bold=True, fill=HEADER_FILL)
    hdr(ws, 1, len(headers))

    keys_in_order = [
        ("Net sales", "revenue", "#,##0"),
        ("COGS", "cogs", "#,##0"),
        ("Gross profit", "gross_profit", "#,##0"),
        ("Gross margin", "gm_pct", "0.0%"),
        ("SG&A", "sga", "#,##0"),
        ("Operating income", "operating_income", "#,##0"),
        ("Operating margin", "op_pct", "0.0%"),
        ("Net income", "net_income", "#,##0"),
        ("Net margin", "ni_pct", "0.0%"),
        ("Diluted EPS", "eps_diluted", "$0.00"),
        ("Diluted shares (M)", "shares_m", "#,##0.0"),
        ("CFO", "cfo", "#,##0"),
        ("Capex", "capex", "#,##0"),
        ("FCF (CFO - Capex)", "fcf", "#,##0"),
        ("Cash & equivalents", "cash", "#,##0"),
        ("Inventory", "inventory", "#,##0"),
        ("Accounts receivable", "receivables", "#,##0"),
        ("Total assets", "total_assets", "#,##0"),
        ("Stockholders' equity", "stockholders_equity", "#,##0"),
    ]
    row = 2
    for label, col, fmt in keys_in_order:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT if "margin" in label.lower() or label.startswith(("Diluted EPS", "Net sales", "Gross profit", "Operating income", "Net income", "FCF")) else BODY_FONT
        for j, ar in enumerate(annual_rows, 2):
            rev = float(ar.get("revenue") or 0)
            if col == "gm_pct":
                v = float(ar.get("gross_profit") or 0) / rev if rev else None
            elif col == "op_pct":
                v = float(ar.get("operating_income") or 0) / rev if rev else None
            elif col == "ni_pct":
                v = float(ar.get("net_income") or 0) / rev if rev else None
            elif col == "shares_m":
                sd = ar.get("shares_diluted")
                v = float(sd) / 1_000_000 if sd else None
            elif col == "fcf":
                v = float(ar.get("cfo") or 0) - float(ar.get("capex") or 0)
            else:
                raw = ar.get(col)
                v = float(raw) if raw not in (None, "") else None
            cell = ws.cell(row=row, column=j, value=v)
            cell.number_format = fmt
            cell.font = BODY_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Source: extracted/annual.csv (XBRL companyfacts) -- backed by 10-Ks in raw/filings/").font = Font(size=9, italic=True, color="666666")

    # Segment sales matrix
    row += 2
    write(ws, row, ["Net sales by segment x brand ($000s)"], bold=True, fill=SUBHEADER_FILL)
    row += 1
    write(ws, row, ["", "FY2022A", "FY2023A", "FY2024A", "FY2025A"], bold=True, fill=SUBHEADER_FILL)
    row += 1
    for seg in ["Americas", "International"]:
        for brand in ["Vita Coco Coconut Water", "Private Label", "Other"]:
            ws.cell(row=row, column=1, value=f"  {seg} - {brand}").font = BODY_FONT
            for j, y in enumerate(["2022", "2023", "2024", "2025"], 2):
                v = fy_seg.get((y, seg, brand))
                cell = ws.cell(row=row, column=j, value=v)
                cell.number_format = "#,##0"
                cell.font = BODY_FONT
            row += 1
        ws.cell(row=row, column=1, value=f"  {seg} subtotal").font = BOLD_FONT
        for j, y in enumerate(["2022", "2023", "2024", "2025"], 2):
            tot = sum(fy_seg.get((y, seg, b)) or 0 for b in ["Vita Coco Coconut Water", "Private Label", "Other"])
            c = ws.cell(row=row, column=j, value=tot)
            c.number_format = "#,##0"
            c.font = BOLD_FONT
            c.fill = TOTAL_FILL
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Source: extracted/segment_history.csv (parsed from 8-K Ex 99.1 press releases)").font = Font(size=9, italic=True, color="666666")


# ---------------------------------------------------------------------------
# Income Statement
# ---------------------------------------------------------------------------
def build_is(wb: Workbook, fy_seg, ar) -> None:
    ws = wb.create_sheet("IS")
    cw(ws, [42, 14, 14, 14, 14])
    write(ws, 1, ["INCOME STATEMENT ($000s)", "FY2025A", "FY2026E", "FY2027E", "FY2028E"], bold=True, fill=HEADER_FILL)
    hdr(ws, 1, 5)

    rev_lines = [
        ("Americas - Vita Coco Coconut Water", FY25["am_vc"],  "Americas - Vita Coco Coconut Water (% YoY)"),
        ("Americas - Private Label",            FY25["am_pl"],  "Americas - Private Label (% YoY)"),
        ("Americas - Other",                    FY25["am_oth"], "Americas - Other (% YoY)"),
        ("International - Vita Coco Coconut Water", FY25["intl_vc"],  "International - Vita Coco Coconut Water (% YoY)"),
        ("International - Private Label",       FY25["intl_pl"], "International - Private Label (% YoY)"),
        ("International - Other",               FY25["intl_oth"], "International - Other (% YoY)"),
    ]
    for i, (label, val, key) in enumerate(rev_lines):
        r = i + 2
        ws.cell(row=r, column=1, value=label).font = BODY_FONT
        ws.cell(row=r, column=2, value=val).number_format = "#,##0"
        ws.cell(row=r, column=2).font = BODY_FONT
        live_row = ar[key]
        for j, live_col in [(3, "K"), (4, "L"), (5, "M")]:
            prev = get_column_letter(j - 1)
            ws.cell(row=r, column=j, value=f"={prev}{r}*(1+Assumptions!{live_col}{live_row})").number_format = "#,##0"
            ws.cell(row=r, column=j).font = BODY_FONT

    # Row 8: Total net sales
    write(ws, 8, ["  Net sales (total)"], bold=True, fill=TOTAL_FILL)
    for col in range(2, 6):
        cl = get_column_letter(col)
        c = ws.cell(row=8, column=col, value=f"=SUM({cl}2:{cl}7)")
        c.number_format = "#,##0"
        c.font = BOLD_FONT
        c.fill = TOTAL_FILL

    # Row 10: Americas subtotal
    ws.cell(row=10, column=1, value="  Americas subtotal").font = BODY_FONT
    for col in range(2, 6):
        cl = get_column_letter(col)
        ws.cell(row=10, column=col, value=f"=SUM({cl}2:{cl}4)").number_format = "#,##0"
    # Row 11: International subtotal
    ws.cell(row=11, column=1, value="  International subtotal").font = BODY_FONT
    for col in range(2, 6):
        cl = get_column_letter(col)
        ws.cell(row=11, column=col, value=f"=SUM({cl}5:{cl}7)").number_format = "#,##0"

    # Row 12-13: GP by segment
    am_gm_row = ar["Americas gross margin (%)"]
    intl_gm_row = ar["International gross margin (%)"]
    ws.cell(row=12, column=1, value="  GP - Americas").font = BODY_FONT
    ws.cell(row=12, column=2, value=FY25["gp_americas"]).number_format = "#,##0"
    ws.cell(row=13, column=1, value="  GP - International").font = BODY_FONT
    ws.cell(row=13, column=2, value=FY25["gp_intl"]).number_format = "#,##0"
    for j, live_col in [(3, "K"), (4, "L"), (5, "M")]:
        cl = get_column_letter(j)
        ws.cell(row=12, column=j, value=f"={cl}10*Assumptions!{live_col}{am_gm_row}").number_format = "#,##0"
        ws.cell(row=13, column=j, value=f"={cl}11*Assumptions!{live_col}{intl_gm_row}").number_format = "#,##0"

    # Row 14: Gross profit
    ws.cell(row=14, column=1, value="Gross profit").font = BOLD_FONT
    ws.cell(row=14, column=1).fill = TOTAL_FILL
    for col in range(2, 6):
        cl = get_column_letter(col)
        c = ws.cell(row=14, column=col, value=f"={cl}12+{cl}13")
        c.number_format = "#,##0"; c.font = BOLD_FONT; c.fill = TOTAL_FILL

    # Row 15: GM%
    ws.cell(row=15, column=1, value="  Gross margin").font = BODY_FONT
    for col in range(2, 6):
        cl = get_column_letter(col)
        ws.cell(row=15, column=col, value=f"={cl}14/{cl}8").number_format = "0.0%"

    # Row 17: SG&A
    sga_row = ar["SG&A as % of net sales"]
    ws.cell(row=17, column=1, value="SG&A").font = BODY_FONT
    ws.cell(row=17, column=2, value=FY25["sga"]).number_format = "#,##0"
    for j, live_col in [(3, "K"), (4, "L"), (5, "M")]:
        cl = get_column_letter(j)
        ws.cell(row=17, column=j, value=f"={cl}8*Assumptions!{live_col}{sga_row}").number_format = "#,##0"

    # Row 18: Op income
    ws.cell(row=18, column=1, value="Operating income").font = BOLD_FONT
    for col in range(2, 6):
        cl = get_column_letter(col)
        c = ws.cell(row=18, column=col, value=f"={cl}14-{cl}17")
        c.number_format = "#,##0"; c.font = BOLD_FONT
    # Row 19: op margin
    ws.cell(row=19, column=1, value="  Operating margin").font = BODY_FONT
    for col in range(2, 6):
        cl = get_column_letter(col)
        ws.cell(row=19, column=col, value=f"={cl}18/{cl}8").number_format = "0.0%"

    # Row 20: Non-operating income (FX, interest, etc.)
    nonop_row = ar["Non-operating income ($M)"]
    ws.cell(row=20, column=1, value="Non-operating income (FX/int.)").font = BODY_FONT
    ws.cell(row=20, column=2, value=FY25["non_op"]).number_format = "#,##0"
    for j, live_col in [(3, "K"), (4, "L"), (5, "M")]:
        cl = get_column_letter(j)
        ws.cell(row=20, column=j, value=f"=Assumptions!{live_col}{nonop_row}*1000").number_format = "#,##0"  # convert $M to $000s

    # Row 21: pretax
    ws.cell(row=21, column=1, value="Pretax income").font = BODY_FONT
    for col in range(2, 6):
        cl = get_column_letter(col)
        ws.cell(row=21, column=col, value=f"={cl}18+{cl}20").number_format = "#,##0"

    # Row 22: tax
    tax_row = ar["Effective tax rate (%)"]
    ws.cell(row=22, column=1, value="  Income tax").font = BODY_FONT
    ws.cell(row=22, column=2, value=FY25["tax"]).number_format = "#,##0"
    for j, live_col in [(3, "K"), (4, "L"), (5, "M")]:
        cl = get_column_letter(j)
        ws.cell(row=22, column=j, value=f"={cl}21*Assumptions!{live_col}{tax_row}").number_format = "#,##0"

    # Row 23: NI
    ws.cell(row=23, column=1, value="Net income").font = BOLD_FONT
    ws.cell(row=23, column=1).fill = TOTAL_FILL
    for col in range(2, 6):
        cl = get_column_letter(col)
        c = ws.cell(row=23, column=col, value=f"={cl}21-{cl}22")
        c.number_format = "#,##0"; c.font = BOLD_FONT; c.fill = TOTAL_FILL

    # Row 24-25: D&A and SBC (memo)
    da_row = ar["D&A as % of net sales"]
    sbc_row = ar["Stock-based comp as % of net sales"]
    ws.cell(row=24, column=1, value="  D&A (memo)").font = BODY_FONT
    ws.cell(row=24, column=2, value=FY25["da"]).number_format = "#,##0"
    ws.cell(row=25, column=1, value="  Stock-based comp (memo)").font = BODY_FONT
    ws.cell(row=25, column=2, value=FY25["sbc"]).number_format = "#,##0"
    for j, live_col in [(3, "K"), (4, "L"), (5, "M")]:
        cl = get_column_letter(j)
        ws.cell(row=24, column=j, value=f"={cl}8*Assumptions!{live_col}{da_row}").number_format = "#,##0"
        ws.cell(row=25, column=j, value=f"={cl}8*Assumptions!{live_col}{sbc_row}").number_format = "#,##0"

    # Row 26: Adj EBITDA
    ws.cell(row=26, column=1, value="Adj EBITDA").font = BOLD_FONT
    ws.cell(row=26, column=1).fill = TOTAL_FILL
    for col in range(2, 6):
        cl = get_column_letter(col)
        c = ws.cell(row=26, column=col, value=f"={cl}18+{cl}24+{cl}25")
        c.number_format = "#,##0"; c.font = BOLD_FONT; c.fill = TOTAL_FILL
    ws.cell(row=27, column=1, value="  Adj EBITDA margin").font = BODY_FONT
    for col in range(2, 6):
        cl = get_column_letter(col)
        ws.cell(row=27, column=col, value=f"={cl}26/{cl}8").number_format = "0.0%"

    # Row 29: Diluted shares (M)
    sh_row = ar["Diluted shares outstanding (M)"]
    ws.cell(row=29, column=1, value="  Diluted shares (M)").font = BODY_FONT
    ws.cell(row=29, column=2, value=FY25["shares_dil"]).number_format = "#,##0.00"
    for j, live_col in [(3, "K"), (4, "L"), (5, "M")]:
        cl = get_column_letter(j)
        ws.cell(row=29, column=j, value=f"=Assumptions!{live_col}{sh_row}").number_format = "#,##0.00"

    # Row 31: Diluted EPS
    ws.cell(row=31, column=1, value="Diluted EPS").font = BOLD_FONT
    for col in range(2, 6):
        cl = get_column_letter(col)
        c = ws.cell(row=31, column=col, value=f"={cl}23/{cl}29/1000")
        c.number_format = "$0.00"; c.font = BOLD_FONT


# ---------------------------------------------------------------------------
# Balance Sheet (ratio-driven, simplified)
# ---------------------------------------------------------------------------
def build_bs(wb: Workbook, ar) -> None:
    ws = wb.create_sheet("BS")
    cw(ws, [42, 14, 14, 14, 14])
    write(ws, 1, ["BALANCE SHEET ($000s)", "FY2025A", "FY2026E", "FY2027E", "FY2028E"], bold=True, fill=HEADER_FILL)
    hdr(ws, 1, 5)

    # Anchors
    cash = FY25["cash"]; rec = FY25["ar"]; inv = FY25["inv"]; tca = FY25["tca"]
    ta = FY25["ta"]; ap = FY25["ap"]; tcl = FY25["tcl"]; tl = FY25["tl"]; se = FY25["se"]
    other_ca = tca - cash - rec - inv
    ppe = ta - tca
    other_cl = tcl - ap
    ncl = tl - tcl

    rows_def = [
        ("Cash & equivalents",  cash),
        ("Accounts receivable", rec),
        ("Inventory",           inv),
        ("Other current assets", other_ca),
        ("  Total current assets", None),
        ("PP&E + intangibles + other", ppe),
        ("TOTAL ASSETS",        None),
        ("Accounts payable",    ap),
        ("Other current liab",  other_cl),
        ("Non-current liab",    ncl),
        ("TOTAL LIABILITIES",   None),
        ("Stockholders' equity", se),
        ("TOTAL LIAB. & EQUITY", None),
        ("Check (Assets - L&E)", None),
    ]
    label_row = {}
    for i, (label, val) in enumerate(rows_def):
        r = i + 2
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT if label.startswith(("TOTAL", "  Total", "Check")) else BODY_FONT
        if val is not None:
            ws.cell(row=r, column=2, value=val).number_format = "#,##0"
        label_row[label] = r

    # Forecast formulas
    for col in range(3, 6):
        cl = get_column_letter(col)
        prev = get_column_letter(col - 1)
        # Working capital items as ratios to FY25 revenue
        ws.cell(row=label_row["Accounts receivable"], column=col, value=f"=IS!{cl}8*{rec/FY25['rev_total']:.6f}").number_format = "#,##0"
        ws.cell(row=label_row["Inventory"],           column=col, value=f"=IS!{cl}8*{inv/FY25['rev_total']:.6f}").number_format = "#,##0"
        ws.cell(row=label_row["Other current assets"], column=col, value=f"=IS!{cl}8*{other_ca/FY25['rev_total']:.6f}").number_format = "#,##0"
        ws.cell(row=label_row["PP&E + intangibles + other"], column=col, value=f"=IS!{cl}8*{ppe/FY25['rev_total']:.6f}").number_format = "#,##0"
        ws.cell(row=label_row["Accounts payable"],    column=col, value=f"=IS!{cl}8*{ap/FY25['rev_total']:.6f}").number_format = "#,##0"
        ws.cell(row=label_row["Other current liab"],  column=col, value=f"=IS!{cl}8*{other_cl/FY25['rev_total']:.6f}").number_format = "#,##0"
        ws.cell(row=label_row["Non-current liab"],    column=col, value=f"=IS!{cl}8*{ncl/FY25['rev_total']:.6f}").number_format = "#,##0"
        # SE roll: prior SE + NI - $15M assumed buybacks
        se_r = label_row["Stockholders' equity"]
        ws.cell(row=se_r, column=col,
                value=f"={prev}{se_r}+IS!{cl}23-15000").number_format = "#,##0"
        # Cash: prior cash + Net Δ cash from CF
        ws.cell(row=label_row["Cash & equivalents"], column=col,
                value=f"={prev}{label_row['Cash & equivalents']}+CF!{cl}11").number_format = "#,##0"

    # Subtotal/total formulas (apply to all columns)
    for col in range(2, 6):
        cl = get_column_letter(col)
        ws.cell(row=label_row["  Total current assets"], column=col,
                value=f"=SUM({cl}{label_row['Cash & equivalents']}:{cl}{label_row['Other current assets']})").number_format = "#,##0"
        ws.cell(row=label_row["TOTAL ASSETS"], column=col,
                value=f"={cl}{label_row['  Total current assets']}+{cl}{label_row['PP&E + intangibles + other']}").number_format = "#,##0"
        ws.cell(row=label_row["TOTAL LIABILITIES"], column=col,
                value=f"=SUM({cl}{label_row['Accounts payable']}:{cl}{label_row['Non-current liab']})").number_format = "#,##0"
        se_r = label_row["Stockholders' equity"]
        ws.cell(row=label_row["TOTAL LIAB. & EQUITY"], column=col,
                value=f"={cl}{label_row['TOTAL LIABILITIES']}+{cl}{se_r}").number_format = "#,##0"
        ws.cell(row=label_row["Check (Assets - L&E)"], column=col,
                value=f"={cl}{label_row['TOTAL ASSETS']}-{cl}{label_row['TOTAL LIAB. & EQUITY']}").number_format = "#,##0"


# ---------------------------------------------------------------------------
# Cash Flow
# ---------------------------------------------------------------------------
def build_cf(wb: Workbook, ar) -> None:
    ws = wb.create_sheet("CF")
    cw(ws, [42, 14, 14, 14, 14])
    write(ws, 1, ["CASH FLOW ($000s)", "FY2025A", "FY2026E", "FY2027E", "FY2028E"], bold=True, fill=HEADER_FILL)
    hdr(ws, 1, 5)

    nwc_row = ar["Net working capital as % of Delta rev"]
    capex_row = ar["Capex as % of net sales"]

    rows_def = [
        ("Net income",          "=IS!B23"),
        ("D&A",                 FY25["da"]),
        ("Stock-based comp",    FY25["sbc"]),
        ("Δ Net working capital", -16500),  # FY25 derived: actual CFO - NI - DA - SBC; rough
        ("CFO",                 None),
        ("Capex",               -FY25["capex"]),
        ("FCF",                 None),
        ("Repurchases",         -11341),  # FY25 actual $11.3M repurchases
        ("Dividends",           0),
        ("Other (FX, etc.)",    0),
        ("Net Δ cash",          None),
    ]
    label_row = {}
    for i, (label, val) in enumerate(rows_def):
        r = i + 2
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT if label in ("CFO", "FCF", "Net Δ cash") else BODY_FONT
        if val is not None:
            cell = ws.cell(row=r, column=2, value=val)
            cell.number_format = "#,##0"
            cell.font = BODY_FONT
        label_row[label] = r

    for col in range(3, 6):
        cl = get_column_letter(col)
        prev = get_column_letter(col - 1)
        ws.cell(row=label_row["Net income"],     column=col, value=f"=IS!{cl}23").number_format = "#,##0"
        ws.cell(row=label_row["D&A"],            column=col, value=f"=IS!{cl}24").number_format = "#,##0"
        ws.cell(row=label_row["Stock-based comp"], column=col, value=f"=IS!{cl}25").number_format = "#,##0"
        ws.cell(row=label_row["Δ Net working capital"], column=col,
                value=f"=-Assumptions!{['','K','L','M','N'][col-2]}{nwc_row}*(IS!{cl}8-IS!{prev}8)").number_format = "#,##0"
        ws.cell(row=label_row["CFO"], column=col,
                value=f"={cl}{label_row['Net income']}+{cl}{label_row['D&A']}+{cl}{label_row['Stock-based comp']}+{cl}{label_row['Δ Net working capital']}").number_format = "#,##0"
        ws.cell(row=label_row["Capex"], column=col,
                value=f"=-IS!{cl}8*Assumptions!{['','K','L','M','N'][col-2]}{capex_row}").number_format = "#,##0"
        ws.cell(row=label_row["FCF"], column=col,
                value=f"={cl}{label_row['CFO']}+{cl}{label_row['Capex']}").number_format = "#,##0"
        ws.cell(row=label_row["Repurchases"], column=col, value=-15000).number_format = "#,##0"
        ws.cell(row=label_row["Dividends"], column=col, value=0).number_format = "#,##0"
        ws.cell(row=label_row["Other (FX, etc.)"], column=col, value=0).number_format = "#,##0"
        ws.cell(row=label_row["Net Δ cash"], column=col,
                value=f"={cl}{label_row['FCF']}+{cl}{label_row['Repurchases']}+{cl}{label_row['Dividends']}+{cl}{label_row['Other (FX, etc.)']}").number_format = "#,##0"

    # Compute FY25A CFO/FCF/NetDeltaCash
    ws.cell(row=label_row["CFO"], column=2,
            value=f"=B{label_row['Net income']}+B{label_row['D&A']}+B{label_row['Stock-based comp']}+B{label_row['Δ Net working capital']}").number_format = "#,##0"
    ws.cell(row=label_row["FCF"], column=2,
            value=f"=B{label_row['CFO']}+B{label_row['Capex']}").number_format = "#,##0"
    ws.cell(row=label_row["Net Δ cash"], column=2,
            value=f"=B{label_row['FCF']}+B{label_row['Repurchases']}+B{label_row['Dividends']}+B{label_row['Other (FX, etc.)']}").number_format = "#,##0"


# ---------------------------------------------------------------------------
# Valuation
# ---------------------------------------------------------------------------
def build_valuation(wb: Workbook, ar) -> None:
    ws = wb.create_sheet("Valuation")
    cw(ws, [44, 16, 16, 16, 16])

    write(ws, 1, ["VALUATION", "FY2026E", "FY2027E", "FY2028E"], bold=True, fill=HEADER_FILL)
    hdr(ws, 1, 4)

    ws.cell(row=2, column=1, value="Adj EBITDA ($M)").font = BODY_FONT
    ws.cell(row=2, column=2, value="=IS!C26/1000").number_format = "#,##0.0"
    ws.cell(row=2, column=3, value="=IS!D26/1000").number_format = "#,##0.0"
    ws.cell(row=2, column=4, value="=IS!E26/1000").number_format = "#,##0.0"

    ws.cell(row=3, column=1, value="Net income ($M)").font = BODY_FONT
    ws.cell(row=3, column=2, value="=IS!C23/1000").number_format = "#,##0.0"
    ws.cell(row=3, column=3, value="=IS!D23/1000").number_format = "#,##0.0"
    ws.cell(row=3, column=4, value="=IS!E23/1000").number_format = "#,##0.0"

    ws.cell(row=4, column=1, value="FCF ($M)").font = BODY_FONT
    ws.cell(row=4, column=2, value="=CF!C8/1000").number_format = "#,##0.0"
    ws.cell(row=4, column=3, value="=CF!D8/1000").number_format = "#,##0.0"
    ws.cell(row=4, column=4, value="=CF!E8/1000").number_format = "#,##0.0"

    write(ws, 6, ["EXIT EV/EBITDA APPROACH (anchor)"], bold=True, fill=SUBHEADER_FILL)

    em_row = ar["Exit EV/EBITDA multiple (FY27)"]
    cash_row = ar["Cash at FY25 ($M)"]
    debt_row = ar["Debt at FY25 ($M)"]
    wacc_row = ar["WACC (%)"]
    sh_row = ar["Diluted shares outstanding (M)"]

    ws.cell(row=7, column=1, value="Exit multiple on FY27E EBITDA").font = BODY_FONT
    ws.cell(row=7, column=2, value=f"=Assumptions!L{em_row}").number_format = "0.0\"x\""
    ws.cell(row=8, column=1, value="Implied EV at end of FY27 ($M)").font = BODY_FONT
    ws.cell(row=8, column=2, value=f"=C2*Assumptions!L{em_row}").number_format = "#,##0.0"
    ws.cell(row=9, column=1, value="(+) Cash at end of FY27 ($M)").font = BODY_FONT
    # Cash at FY27 end = FY25 cash + sum FCF FY26+FY27 - sum Repurchases (15+15)
    ws.cell(row=9, column=2,
            value=f"=Assumptions!L{cash_row}+B4+C4-15-15").number_format = "#,##0.0"
    ws.cell(row=10, column=1, value="(-) Debt ($M)").font = BODY_FONT
    ws.cell(row=10, column=2, value=f"=Assumptions!L{debt_row}").number_format = "#,##0.0"
    ws.cell(row=11, column=1, value="Equity value at end of FY27 ($M)").font = BOLD_FONT
    ws.cell(row=11, column=2, value="=B8+B9-B10").number_format = "#,##0.0"
    ws.cell(row=11, column=2).fill = TOTAL_FILL
    ws.cell(row=12, column=1, value="Diluted shares (M)").font = BODY_FONT
    ws.cell(row=12, column=2, value=f"=Assumptions!L{sh_row}").number_format = "#,##0.00"
    ws.cell(row=13, column=1, value="Implied price at end of FY27 ($)").font = BOLD_FONT
    ws.cell(row=13, column=2, value="=B11/B12").number_format = "$0.00"

    ws.cell(row=15, column=1, value="WACC (for discount)").font = BODY_FONT
    ws.cell(row=15, column=2, value=f"=Assumptions!L{wacc_row}").number_format = "0.0%"
    ws.cell(row=16, column=1, value="Discount window: end FY27 -> Apr 2027 (~8 months)").font = BODY_FONT

    write(ws, 18, ["12-MONTH PRICE TARGET (LIVE)"], bold=True, fill=SUBHEADER_FILL)
    ws.cell(row=19, column=1, value="Approach: Equity value at end FY27 / shares, discount to Apr 2027 (12m fwd)").font = BODY_FONT
    ws.cell(row=20, column=1, value="Discount factor = 1/(1+WACC)^0.67").font = BODY_FONT
    ws.cell(row=21, column=1, value="Discounted equity value ($M)").font = BODY_FONT
    ws.cell(row=21, column=2, value="=B11/(1+B15)^0.67").number_format = "#,##0.0"
    ws.cell(row=22, column=1, value="12-MONTH PRICE TARGET ($)").font = BOLD_FONT
    ws.cell(row=22, column=1).fill = TOTAL_FILL
    ws.cell(row=22, column=2, value="=B21/B12").number_format = "$0.00"
    ws.cell(row=22, column=2).fill = TOTAL_FILL
    ws.cell(row=22, column=2).font = Font(size=12, bold=True, color="2F5496")

    # Cross-check with P/E
    write(ws, 24, ["CROSS-CHECK: P/E"], bold=True, fill=SUBHEADER_FILL)
    ws.cell(row=25, column=1, value="FY27E EPS ($)").font = BODY_FONT
    ws.cell(row=25, column=2, value="=IS!D31").number_format = "$0.00"
    ws.cell(row=26, column=1, value="P/E target on FY27 EPS").font = BODY_FONT
    ws.cell(row=26, column=2, value=30.0).number_format = "0.0"
    ws.cell(row=27, column=1, value="Implied price at end FY27 ($)").font = BODY_FONT
    ws.cell(row=27, column=2, value="=B25*B26").number_format = "$0.00"
    ws.cell(row=28, column=1, value="Discounted to Apr 2027 (12m fwd) at WACC").font = BOLD_FONT
    ws.cell(row=28, column=2, value="=B27/(1+B15)^0.67").number_format = "$0.00"

    # FCF / DCF cross-check (terminal value Gordon growth on FY28 FCF)
    write(ws, 30, ["CROSS-CHECK: DCF (Gordon growth on FY28 FCF)"], bold=True, fill=SUBHEADER_FILL)
    g_row = ar["Terminal growth rate (%)"]
    ws.cell(row=31, column=1, value="FY28E FCF ($M)").font = BODY_FONT
    ws.cell(row=31, column=2, value="=D4").number_format = "#,##0.0"
    ws.cell(row=32, column=1, value="WACC (%)").font = BODY_FONT
    ws.cell(row=32, column=2, value="=B15").number_format = "0.0%"
    ws.cell(row=33, column=1, value="Terminal growth (%)").font = BODY_FONT
    ws.cell(row=33, column=2, value=f"=Assumptions!L{g_row}").number_format = "0.0%"
    ws.cell(row=34, column=1, value="Terminal value at end FY28 ($M)").font = BODY_FONT
    ws.cell(row=34, column=2, value="=B31*(1+B33)/(B32-B33)").number_format = "#,##0.0"
    ws.cell(row=35, column=1, value="PV of FY26-28 FCF ($M)").font = BODY_FONT
    # Discount FY26 by 0.5 yr (mid-year), FY27 by 1.5, FY28 by 2.5 (mid-year of each fiscal year)
    ws.cell(row=35, column=2, value="=B4/(1+B15)^0.5+C4/(1+B15)^1.5+D4/(1+B15)^2.5").number_format = "#,##0.0"
    ws.cell(row=36, column=1, value="PV of terminal value ($M)").font = BODY_FONT
    ws.cell(row=36, column=2, value="=B34/(1+B15)^3.0").number_format = "#,##0.0"
    ws.cell(row=37, column=1, value="Enterprise value (DCF) ($M)").font = BOLD_FONT
    ws.cell(row=37, column=2, value="=B35+B36").number_format = "#,##0.0"
    ws.cell(row=38, column=1, value="(+) Cash now ($M)").font = BODY_FONT
    ws.cell(row=38, column=2, value=f"=Assumptions!L{cash_row}").number_format = "#,##0.0"
    ws.cell(row=39, column=1, value="Equity value (DCF) ($M)").font = BOLD_FONT
    ws.cell(row=39, column=2, value="=B37+B38").number_format = "#,##0.0"
    ws.cell(row=40, column=1, value="DCF price per share ($)").font = BOLD_FONT
    ws.cell(row=40, column=2, value="=B39/B12").number_format = "$0.00"


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    annual, fy_seg = load_history()
    wb = Workbook()
    wb.remove(wb.active)

    build_cover(wb)  # creates Cover at index 0 (creates blank Cover for Scenario named range)
    rowmap = build_assumptions(wb)
    build_history(wb, annual, fy_seg)
    build_is(wb, fy_seg, rowmap)
    build_bs(wb, rowmap)
    build_cf(wb, rowmap)
    build_valuation(wb, rowmap)
    wb.move_sheet("Cover", offset=-len(wb.sheetnames))

    wb.save(OUT)
    print(f"OK wrote {OUT.relative_to(REPO).as_posix()}")


if __name__ == "__main__":
    main()
